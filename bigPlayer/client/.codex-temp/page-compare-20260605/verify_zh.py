from openpyxl import load_workbook
p = r"C:\Users\Administrator\AppData\Roaming\Code\User\project manage\bigPlayer\client\.codex-temp\page-compare-20260605\bigplayer_home_compare_report_zh.xlsx"
wb = load_workbook(p)
ws = wb["Gap_List_CN"]
v = ws['C2'].value
print(wb.sheetnames)
print(ws.max_row, ws.max_column)
print(v)
print([ord(ch) for ch in v[:5]])
