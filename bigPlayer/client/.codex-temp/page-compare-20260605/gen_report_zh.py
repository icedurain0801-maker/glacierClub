from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

base = Path(r"C:\Users\Administrator\AppData\Roaming\Code\User\project manage\bigPlayer\client\.codex-temp\page-compare-20260605")
out = base / "bigplayer_home_compare_report_zh.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Gap_List_CN"

headers = ["序号", "维度", "不达标项", "需求页表现（home-en）", "实现页表现（club）", "影响级别", "修复建议", "证据截图（文件）"]
rows = [
    [1, "页面框架", "桌面端整体构图不一致", "主体内容在手机机模内居中展示，外层保留明显留白。", "页面直接铺满内容流，缺少机模容器与外层留白。", "高", "增加桌面机模壳层与居中容器，内容在机模内滚动。", "requirement-desktop.png vs implementation-desktop.png"],
    [2, "页面框架", "右上模式切换与主题按钮缺失", "右上有版本切换胶囊（网页版/AP版/小程序版/简洁版）和主题按钮。", "未看到对应控件。", "高", "补齐右上控件组，对齐间距、圆角、阴影和层级。", "requirement-desktop.png / requirement-mobile.png vs implementation-desktop.png / implementation-mobile.png"],
    [3, "新手引导", "首屏引导弹层未实现", "首屏有 Step 1 of 2 引导卡，含 Skip/Next 与分页点。", "未出现引导弹层与步骤交互。", "高", "实现首访引导浮层、步骤切换与关闭状态持久化。", "requirement-desktop.png / requirement-mobile.png vs implementation-desktop.png / implementation-mobile.png"],
    [4, "内容模块", "主推荐模块形态不一致", "Top Recommendations 为大图主卡，含标签、摘要、互动信息。", "对应区域是资讯列表，缺少大图主卡形态。", "高", "首模块改为视觉主卡组件，补齐封面图、标签、摘要和互动信息。", "requirement-mobile.png vs implementation-mobile.png"],
    [5, "卡片样式", "资讯卡视觉语言不一致", "卡片为亮底+封面图+信息遮罩层，层级清晰。", "卡片为深色条块，图片层与信息层分离不足。", "中", "重构资讯卡样式，统一封面图、遮罩层、圆角和间距体系。", "requirement-mobile.png vs implementation-mobile.png"],
    [6, "头部信息架构", "头部结构与语义偏离需求", "头部强调品牌位、搜索和关系入口。", "当前为另一套频道结构，信息层级不同。", "中", "按需求页重排头部 IA：品牌位、搜索入口、关系入口、频道层。", "requirement-mobile.png vs implementation-mobile.png"],
    [7, "底部导航", "底部导航视觉与状态处理不一致", "图标风格轻量，激活态与中心按钮层级明确。", "图标风格、间距、字号和重心与需求差异较大。", "中", "对齐图标资源、激活色、字号、间距和中心按钮尺寸。", "requirement-mobile.png vs implementation-mobile.png"],
    [8, "响应式节奏", "移动端首屏留白和分组节奏不足", "首屏在引导、Banner、主推荐之间有明确节奏。", "模块更紧凑，分组留白不足。", "中", "重设移动端垂直间距：区块间距、标题与内容间距、卡片间距。", "requirement-mobile.png vs implementation-mobile.png"],
    [9, "布局稳定性", "底部固定栏对内容有压迫感", "底栏与内容过渡自然，底部内容不受干扰。", "底部内容与固定栏距离过近，可读性受影响。", "中", "内容容器增加底部安全区内边距（safe area + nav height）。", "implementation-mobile.png"],
    [10, "国际化", "语言版本与需求目标不一致", "需求页为英文海外版本。", "当前实现链接为中文文案（zh-CN）。", "中", "补齐英文资源并验证 `lang=en` 与默认语言回退策略。", "requirement-desktop.png vs implementation-desktop.png"],
]

ws.append(headers)
for r in rows:
    ws.append(r)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9D9D9")

for c in ws[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

sev_col = headers.index("影响级别") + 1
for i in range(2, ws.max_row + 1):
    c = ws.cell(i, sev_col)
    if c.value == "高":
        c.fill = PatternFill("solid", fgColor="F8CBAD")
        c.font = Font(color="9C0006", bold=True)
    elif c.value == "中":
        c.fill = PatternFill("solid", fgColor="FFE699")
        c.font = Font(color="7F6000", bold=True)

widths = {1: 8, 2: 14, 3: 26, 4: 42, 5: 42, 6: 10, 7: 44, 8: 60}
for idx, w in widths.items():
    ws.column_dimensions[get_column_letter(idx)].width = w
ws.freeze_panes = "A2"

meta = wb.create_sheet("Notes_CN")
meta_rows = [
    ("报告名称", "BigPlayer 海外首页对比不达标清单"),
    ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("需求页", "https://icedurain0801-maker.github.io/glacierClub/bigPlayer/client/overseas/home/home-en.html"),
    ("实现页", "http://club.kubedev.q1.com/?env=web&lang=zh-CN"),
    ("截图目录", str(base)),
    ("备注", "已排除人脸模糊区域影响。"),
]
for k, v in meta_rows:
    meta.append([k, v])

meta.column_dimensions["A"].width = 16
meta.column_dimensions["B"].width = 120
for r in range(1, len(meta_rows) + 1):
    meta.cell(r, 1).font = Font(bold=True)
    meta.cell(r, 1).fill = PatternFill("solid", fgColor="D9E1F2")
    meta.cell(r, 1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    meta.cell(r, 2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    meta.cell(r, 2).alignment = Alignment(vertical="top", wrap_text=True)

wb.save(out)
print(out)
